# _*_ coding: utf-8 _*_
import os
from openpyxl import load_workbook


class ReadExcel:
    file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Data\\testcase.xlsx")

    def __init__(self):
        # 打开已有表格
        wb = load_workbook(filename=self.file_path)
        # 选取当前表格活跃的sheet
        self.ws = wb.active

    def excel_content(self, row, column):
        """
        返回excel中第row行第column列单元格内数据
        :param row:
        :param column:
        :return:
        """
        data = self.ws[row][column-1].value
        return data


# first_row = self.ws[row] 
# first_row_list = []
# for i in range(len(first_row)):
#     print(first_row[i].value)
#     first_row_list.append(first_row[i].value)
# print(first_row_list)


# if __name__ == '__main__':
#     print(ReadExcel().excel_content(2, 3))
